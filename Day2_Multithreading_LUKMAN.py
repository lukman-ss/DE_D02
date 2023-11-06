import os
import requests
import openpyxl
# import PyPDF2 
import re
from PyPDF2 import PdfReader
import time

def count_elapsed_time(start_time, end_time):
    return end_time - start_time
'''
def search_pdf(pdf_path, search_texts):
    found_sentences = {search_text: [] for search_text in search_texts}

    with open(pdf_path, 'rb') as file:
        pdf_reader = PdfReader(file)

        for page in pdf_reader.pages:
            page_text = page.extract_text()

            for search_text in search_texts:
                if search_text in page_text:
                    sentences = page_text.split('\n\n')  # Split paragraphs based on double newlines
                    for sentence in sentences:
                        if search_text in sentence:
                            found_sentences[search_text].append(sentence.strip())

    return found_sentences
'''
def search_pdf(pdf_path, search_texts):
    found_sentences = {search_text: [] for search_text in search_texts}

    try:
        with open(pdf_path, 'rb') as file:
            pdf_reader = PdfReader(file)

            for page in pdf_reader.pages:
                try:
                    page_text = page.extract_text()
                except Exception as e:
                    print(f"Error extracting text from page {page.page_number}: {e}")
                    continue

                for search_text in search_texts:
                    if search_text in page_text:
                        sentences = page_text.split('\n\n')  # Split paragraphs based on double newlines
                        for sentence in sentences:
                            if search_text in sentence:
                                found_sentences[search_text].append(sentence.strip())

        return found_sentences

    except Exception as e:
        print(f"Error reading PDF file: {e}")
        return None
# Usage example:
# pdf_path = "sample.pdf"  # Replace with the path to your PDF file
# search_texts = ["example", "text", "search"]  # Replace with the array of search texts

# found_sentences = search_pdf(pdf_path, search_texts)

# for search_text, sentences in found_sentences.items():
#     if sentences:
#         print(f"Search text '{search_text}' found in the following sentences:")
#         for sentence in sentences:
#             print(f" - {sentence}")
#     else:
#         print(f"Search text '{search_text}' not found in the PDF.")
# def extract_text_from_pdf(pdf_path, split_pages=None):
#     with open(pdf_path, 'rb') as file:
#         pdf_reader = PyPDF2.PdfFileReader(file)
#         text = ""

#         for page_num in range(pdf_reader.numPages):
#             page = pdf_reader.getPage(page_num)
#             page_text = page.extractText()

#             if split_pages and (page_num + 1) in split_pages:
#                 sentences = re.split(r'\. |\n', page_text)  # Split sentences based on periods followed by a space or newlines
#                 text += '.\n'.join(sentences)  # Join the sentences back together with a period and a newline
#             else:
#                 text += page_text

#     return text

def search_text_in_pdf(pdf_text, search_texts):
    found_pages = {}

    for search_text in search_texts:
        found_pages[search_text] = []

    for page_num, page_text in enumerate(pdf_text.split('\n\n')):
        for search_text in search_texts:
            if search_text in page_text:
                found_pages[search_text].append(page_num + 1)

    return found_pages

def search_pdf2(pdf_path, search_text):
    found_pages = []
    
    with open(pdf_path, 'rb') as file:
        pdf_reader = PyPDF2.PdfFileReader(file)

        for page_num in range(pdf_reader.numPages):
            page = pdf_reader.getPage(page_num)
            page_text = page.extractText()
            if search_text in page_text:
                found_pages.append(page_num)
                sentences = re.split(r'\. |\n', page_text)  # Split sentences based on periods followed by a space or newlines
                found_pages.append(sentences)
    return found_pages

def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(row)

    return data

def read_excel_as_array(file_path):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        data.append(list(row))

    return data

def download_pdf(url, save_folder):
    response = requests.get(url)
    if response.status_code == 200:
        filename = url.split("/")[-1]
        save_path = os.path.join(save_folder, filename)

        if not os.path.exists(save_path):  # Check if the file does not exist
            with open(save_path, 'wb') as file:
                file.write(response.content)
            # print(f"PDF file downloaded successfully and saved as {save_path}")
            return filename;
        else:
            print(f"File {filename} already exists in {save_folder}. Skipping download.")
    else:
        print(f"Failed to download the PDF. Status code: {response.status_code}")

if __name__ == "__main__":
    # single thread
    start_time = time.time()  # Get the current time before the process starts

    file_path = "BATAM.xlsx"
    excel_data = read_excel_as_array(file_path)
    no =0;
    # Print the data
    for row in excel_data:
         
        if no > 0:
            
            pdf_url = row[1] # Replace with the actual URL of the PDF file
            folder_path = "pdfs"  # Replace with the path of the specific folder

            # Create the folder if it doesn't exist
            os.makedirs(folder_path, exist_ok=True)

            file_name = download_pdf(pdf_url, folder_path)
            if isinstance(file_name, str):
                pdf_path = "pdfs/" + file_name  # Replace with the path to your PDF file
                
                # pdf_path = "sample.pdf"  # Replace with the path to your PDF file
                search_texts = ["sanksi", "berlaku sampai dengan"] 

                found_sentences = search_pdf(pdf_path, search_texts)
                if found_sentences:
                    for search_text, sentences in found_sentences.items():
                        if sentences:
                            print(f"|'{search_text}'|")
                            # print(f"Search text '{search_text}' found in the following sentences:")
                            for sentence in sentences:
                                print(f"{sentence}")
                            print(f"|")
        no +=1
    end_time = time.time()  # Get the current time after the process ends

    elapsed_time = count_elapsed_time(start_time, end_time)
    print("Elapsed Time:", elapsed_time, "seconds")
    # Elapsed Time: 412.3088369369507 seconds

'''
# multithread
import concurrent.futures

def process_row(row):
    pdf_url = row[1]  # Replace with the actual URL of the PDF file
    folder_path = "pdfs"  # Replace with the path of the specific folder

    # Create the folder if it doesn't exist
    os.makedirs(folder_path, exist_ok=True)

    file_name = download_pdf(pdf_url, folder_path)
    if isinstance(file_name, str):
        
        pdf_path = os.path.join("pdfs", file_name)  # Replace with the path to your PDF file

        search_texts = ["sanksi", "berlaku sampai dengan"] 
        found_sentences = search_pdf(pdf_path, search_texts)

        for search_text, sentences in found_sentences.items():
            if sentences:
                print(f"|{search_text}|")
                for sentence in sentences:
                    print(f"{sentence}")
                print(f"|")

if __name__ == "__main__":
    start_time = time.time()  # Get the current time before the process starts


    file_path = "BATAM.xlsx"
    excel_data = read_excel_as_array(file_path)

    with concurrent.futures.ThreadPoolExecutor() as executor:
        futures = [executor.submit(process_row, row) for row in excel_data[1:]]

        # Wait for all threads to complete
        concurrent.futures.wait(futures)
        
    end_time = time.time()  # Get the current time after the process ends

    elapsed_time = count_elapsed_time(start_time, end_time)
    print("Elapsed Time:", elapsed_time, "seconds")
    # Elapsed Time: 223.83249616622925 seconds
'''