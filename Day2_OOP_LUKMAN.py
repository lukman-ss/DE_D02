from abc import ABC, abstractmethod
from openpyxl import *
# import gspread 
# from oauth2client.service_account import ServiceAccountCredentials
from datetime import datetime
# import gspread
# from oauth2client.service_account import ServiceAccountCredentials
class SavingsAccount(ABC):
    def __init__(self, account_number, account_holder, balance, interest_rate,  username, password):
        self.account_number = account_number
        self.account_holder = account_holder
        self.balance = balance
        self.interest_rate = interest_rate
        self.transaction_history = []
        self.username = username
        self.password = password
        
    @abstractmethod
    def deposit(self, amount):
        self.balance += amount
        self.add_to_transaction_history("Deposit: +{amount}")
        return self.balance

    @abstractmethod
    def withdraw(self, amount):
        if self.balance >= amount:
            self.balance -= amount
            self.add_to_transaction_history("Withdrawal: -{amount}")
            return self.balance
        else:
            print("Insufficient balance.")

    def calculate_interest(self):
        interest = self.balance * (self.interest_rate / 100)
        self.add_to_transaction_history("Interest: +{interest}")
        return interest

    def add_to_transaction_history(self, transaction):
        self.transaction_history.append(transaction)


    def display_transaction_history(self):
        for transaction in self.transaction_history:
            print(transaction)
            
    def display_account_info(self):
        print(f"Account Number: {self.account_number}")
        print(f"Account Holder: {self.account_holder}")
        print(f"Current Balance: {self.balance}")
        
    def export_transaction_history(workbook, sheet_name, transactions):
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(title=sheet_name)

            # Write the header row
            headers = list(transactions[0].keys())
            worksheet.append(headers)

        # Write the transaction data
        for transaction in transactions:
            row_data = list(transaction.values())
            worksheet.append(row_data)
            
    # def export_transaction_history(self,sheets):
    #         sheets = sheets
    #         workbook = openpyxl.Workbook()
    #         sheet = workbook.sheets
    #         sheet.title = "Transaction History"

    #         # Write transaction history to the Excel sheet
    #         row = 1
    #         for transaction in self.transaction_history:
    #             sheet.cell(row=row, column=1, value=transaction)
    #             row += 1

    #         # Save the Excel file locally
    #         filename = f"{self.account_number}_transaction_history.xlsx"
    #         workbook.save(filename)
    #         print(f"Transaction history saved to {filename}.")
    

    # def has_empty_rows(worksheet):
    #     return len(worksheet.col_values(1)) == 1  # Check if the first column is empty (contains only the header row)

    # def append_row_to_google_sheet(sheet_name, data):
    #     # Replace 'credentials.json' with the path to your JSON credentials file
    #     scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    #     credentials = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)

    #     # Replace 'Your Google Sheet Name' with the name of your Google Sheet
    #     gc = gspread.authorize(credentials)
    #     spreadsheet = gc.open('Your Google Sheet Name')

    #     # Try to open the specified sheet, create a new one if it doesn't exist
    #     try:
    #         worksheet = spreadsheet.worksheet(sheet_name)
    #     except gspread.WorksheetNotFound:
    #         worksheet = spreadsheet.add_worksheet(title=sheet_name, rows=1, cols=len(data))
    #         worksheet.append_row([''] * len(data))  # Append an empty row as the header

    #     # Check if the sheet has empty rows before appending
    #     if has_empty_rows(worksheet):
    #         worksheet.append_row(data)
    #     else:
    #         # If the sheet has empty rows, you may perform some other action or handle the case accordingly.
    #         # For example, you might want to log a warning or raise an exception.
    #         print(f"Sheet '{sheet_name}' has empty rows. Skipping append operation.")

    # Usage example:
    # sheet_name = "UniqueSheet"  # Replace with the name of the sheet you want to append to or create
    # data_to_append = [1, 2, 3, 4]  # Replace with the data you want to append as a list

    # append_row_to_google_sheet(sheet_name, data_to_append)
    # def __str__(self):
    #     return f"Account Number: {self.account_number}\nAccount Holder: {self.account_holder}\nBalance: {self.balance}\nInterest Rate: {self.interest_rate}"

class ConventionalSavings(SavingsAccount):
    def __init__(self, account_number, account_holder, balance, interest_rate, min_balance,  username, password):
        super().__init__(account_number, account_holder, balance, interest_rate,username, password )
        self.min_balance = min_balance
        self.username = username
        self.password = password
        self.balance = balance  
        self.account_number = account_number  
        self.username = username  
        
    def add_to_transaction_history(self, transactions):
        # Implement adding transactions to transaction history for conventional savings account
        self.transaction_history.extend(transactions)
    def deposit(self, amount):
        self.balance += amount
        self.add_to_transaction_history("Deposit: +{amount}")
        return self.balance

    def withdraw(self, amount):
        if self.balance - amount >= self.min_balance:
            self.balance -= amount
            self.add_to_transaction_history("Withdrawal: -{amount}")
            return self.balance
        else:
            print("Withdrawal not allowed. Minimum balance requirement not met.")

    def calculate_interest(self):
        interest = self.balance * (self.interest_rate / 100)
        self.balance += interest
        self.add_to_transaction_history("Interest: +{interest}")
        return self.balance
    
    def display_transaction_history(self):
            for transaction in self.transaction_history:
                print(transaction)
                # self, account_number, account_holder, balance, interest_rate,  username, password
class ShariaSavings(SavingsAccount):
    def __init__( self,  account_number, account_holder, balance,interest_rate,username, password):
        super().__init__(account_number, account_holder, balance,interest_rate,username, password)
        self.account_number = account_number
        self.account_holder = account_holder
        self.balance = balance
        self.username = username
        self.username = username
        self.password = password
        
    def calculate_interest(self):
        # Sharia-compliant interest calculation (interest_rate = 0%)
        # self.add_to_transaction_history("Interest: +0")
        return 0
    def deposit(self, amount):
        self.balance += amount
        self.add_to_transaction_history("Deposit: +{amount}")
        return self.balance

    def withdraw(self, amount):
        if amount <= self.balance:
            self.balance -= amount
            self.add_to_transaction_history("Withdrawal: -{amount}")
            return self.balance
            
        else:
            print("Insufficient balance.")

    def calculate_interest(self):
        interest = self.balance * (self.interest_rate / 100)
        self.balance += interest
        self.add_to_transaction_history("Interest: +{interest}")
        return self.balance
    
class Bank:
    def __init__(self):
        self.accounts = []
      
    def register_account(self, account):
        self.accounts.append(account)

    def login(self, username, password):
        for acc in self.accounts:
            if acc.username == username and acc.password == password:
                print("Login successful.")
                return acc
        print("Invalid username or password. Login failed.")
        return None

    def display_all_accounts(self):
        for acc in self.accounts:
            print("---------------------------")
            acc.display_account_info()

def write_xlsx(workbook, sheet_name, new_transactions):
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(title=sheet_name)

        # Write the header row
        headers = list(new_transactions[0].keys())
        worksheet.append(headers)

    # Write the new transaction data
    for transaction in new_transactions:
        row_data = list(transaction.values())
        worksheet.append(row_data)
# Day2: OOP
# key
import gspread
from google.oauth2.service_account import Credentials

def append_data_to_sheet(sheet_name, new_data):
    # Set up the credentials
    creds = Credentials.from_service_account_file('google secret.json', scopes=['https://www.googleapis.com/auth/spreadsheets'])

    # Authenticate and authorize the client
    client = gspread.authorize(creds)

    # Open the Google Sheets file by key
    spreadsheet_key = '1ojN5IITl16HRFvXFrKuEJQOzpjr5L-vu2dZQ_lKTVJg'
    spreadsheet = client.open_by_key(spreadsheet_key)
    worksheet = spreadsheet.worksheet(sheet_name)

    # Find the next empty row
    next_empty_row = len(worksheet.get_all_values()) + 1

    # Append each dictionary as a new row in the worksheet
    for data_dict in new_data:
        cell_data = list(data_dict.values())
        worksheet.append_row(cell_data, value_input_option='USER_ENTERED')
        
    
if __name__ == "__main__":
    current_datetime = datetime.now()
    formatted_datetime = current_datetime.strftime("%Y-%m-%d %H:%M:%S")    
    bank = Bank()
    # Creating accounts
    # account 1
    account1_account_number = "C0001"
    account1_account_holder = "Don Angelo Lukman"
    account1_balance = 1500
    account1_username = 1500
    account1_username = "C0001"
    account1_password = "12345"
    account1 = ConventionalSavings(
        account_number=account1_account_number, 
        account_holder=account1_account_holder,
        balance=account1_balance, interest_rate=3.0, min_balance=1000,username="C0001", password="12345")
    new_transactions =  [{"date": formatted_datetime, "account_number": account1_account_number, "account_name": account1_account_holder, "balance": account1_balance, "student": "Lukman" }]

    sheet_name = "Conventional"  # Replace with the desired sheet name

    workbook = load_workbook('BANK_MANAGEMENT_SYSTEM.xlsx')

    write_xlsx(workbook, sheet_name, new_transactions)

    # Save the updated workbook to a file
    workbook.save('BANK_MANAGEMENT_SYSTEM.xlsx')
    append_data_to_sheet(sheet_name, new_transactions)

    # Registering accounts in the bank
    bank.register_account(account1)
    username_to_login =  account1_username
    password_to_login = account1_password 
    logged_in_account = bank.login(username_to_login, password_to_login)
    if logged_in_account:
        print("Logged in Account:")
        logged_in_account.display_account_info()
    else:
        print("Login failed.")
    # Displaying all accounts in the bank
    print("\nAll Accounts in the Bank:")
    bank.display_all_accounts()
    # account 1 implement deposit
    new_balance = account1.deposit(5000)
    new_transactions =  [{"date": formatted_datetime, "account_number": account1_account_number, "account_name": account1_account_holder, "balance": new_balance, "student": "Lukman" }]
    sheet_name = "Transaction"  # Replace with the desired sheet name
    workbook = load_workbook('BANK_MANAGEMENT_SYSTEM.xlsx')
    write_xlsx(workbook, sheet_name, new_transactions)
    workbook.save('BANK_MANAGEMENT_SYSTEM.xlsx')
    append_data_to_sheet(sheet_name, new_transactions)
    
    
    new_balance = account1.withdraw(1000)
    new_transactions =  [{"date": formatted_datetime, "account_number": account1_account_number, "account_name": account1_account_holder, "balance": new_balance, "student": "Lukman" }]
    sheet_name = "Transaction"  # Replace with the desired sheet name
    workbook = load_workbook('BANK_MANAGEMENT_SYSTEM.xlsx')
    write_xlsx(workbook, sheet_name, new_transactions)
    workbook.save('BANK_MANAGEMENT_SYSTEM.xlsx')
    append_data_to_sheet(sheet_name, new_transactions)
    
    account1.calculate_interest()
    account1.display_account_info()
    account1.display_transaction_history()
    
    # uncomment this line if using accounts 2
    account2_account_number = "S0001"
    account2_account_holder = "Don Leorio Lukman"
    account2_balance = 1500
    account2_username = 1500
    account2_username = "S0001"
    account2_password = "12345"
    account2 = ShariaSavings(account_number=account2_account_number, account_holder=account2_account_holder, balance=account2_balance, username=account2_username, password=account2_password, interest_rate=0)
    
    new_transactions =  [{"date": formatted_datetime, "account_number": account2_account_number, "account_name": account2_account_holder, "balance": account2_balance, "student": "Lukman" }]
    sheet_name = "Sharia"  # Replace with the desired sheet name
    workbook = load_workbook('BANK_MANAGEMENT_SYSTEM.xlsx')
    write_xlsx(workbook, sheet_name, new_transactions)
    # Save the updated workbook to a file
    workbook.save('BANK_MANAGEMENT_SYSTEM.xlsx')
    append_data_to_sheet(sheet_name, new_transactions)
    
    bank.register_account(account2)
    username_to_login =  account2_username
    password_to_login = account2_password 
    logged_in_account = bank.login(username_to_login, password_to_login)
    if logged_in_account:
        print("Logged in Account:")
        logged_in_account.display_account_info()
    else:
        print("Login failed.")
        
    new_balance = account2.deposit(10000)
    new_transactions =  [{"date": formatted_datetime, "account_number": account2_account_number, "account_name": account2_account_holder, "balance": new_balance, "student": "Lukman" }]
    sheet_name = "Transaction"  # Replace with the desired sheet name
    workbook = load_workbook('BANK_MANAGEMENT_SYSTEM.xlsx')
    write_xlsx(workbook, sheet_name, new_transactions)
    workbook.save('BANK_MANAGEMENT_SYSTEM.xlsx')
    append_data_to_sheet(sheet_name, new_transactions)

    
    new_balance = account2.withdraw(4000)
    ew_transactions =  [{"date": formatted_datetime, "account_number": account2_account_number, "account_name": account2_account_holder, "balance": new_balance, "student": "Lukman" }]
    sheet_name = "Transaction"  # Replace with the desired sheet name
    workbook = load_workbook('BANK_MANAGEMENT_SYSTEM.xlsx')
    write_xlsx(workbook, sheet_name, new_transactions)
    workbook.save('BANK_MANAGEMENT_SYSTEM.xlsx')
    append_data_to_sheet(sheet_name, new_transactions)
    